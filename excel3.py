import openpyxl
from openpyxl import Workbook

file = 'C:\\Users\\Ozgur\\Desktop\\htm\\macro.xlsm'
wb = openpyxl.load_workbook(filename=file,read_only=False,keep_vba=True)
#print (wb.sheetnames)
ws = wb.get_sheet_by_name('rapor')
#ws.cell(row=int(x)*10,column=3,value=104)
name="volvo"
ws.cell(row=1,column=1,value=name)
wb.save(file)
wb.close()
quit()
#print(ws.max_row)
a= ws.cell(row=1,column=2)
#t=a.value
#print(a.value)
#ws.cell(row=1,column=3, value=t)
t=[]
a=[]
for row in ws.iter_rows():
 
    
  for cell in row:
    a.append(cell.value)
  
  t.append(a)	
  a=[]   
print (t[0])
